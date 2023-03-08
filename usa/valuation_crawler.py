# valuation_crawler.py ---------------------------------------------------------------------------------------------
#   Crawls through list of U.S. public companies to carry out fundamental valuation and store the results for later
# sorting and selection of companies for manual re-valuation
# ------------------------------------------------------------------------------------------------------------------
import numpy as np
import pandas as pd
from dcf_valuation import get_statements, get_fundamentals, value_company
from random import randint
from openpyxl import load_workbook
import os
from datetime import date
import time

# Use exchange names to shortlist companies 
# Throw away all the companies that are financial
def valuation_crawler(industry_name = None): 
    indname_df = pd.read_excel('/home/dinesh/Documents/Valuations/adamodaran/indname.xlsx').fillna('') # Extract stocks, industry list
    us_stocks = indname_df[indname_df['Exchange:Ticker'].str.contains("Nasdaq|NYSE")] # Extract a sublist containing stocks listed in US exchanges
    us_stocks.reset_index(inplace=True) # Reset the index to start from 0 and increment sequentially (so they can be selected using a random number generator
   
    if industry_name == None: 
        sel_stocks = us_stocks
    else:
        try:
            sel_stocks = us_stocks[us_stocks['Industry Group'] == industry_name]
        except:
            print("Invalid indistry name") 
    
    sel_stocks.reset_index(inplace=True) # Reset the index to start from 0 and increment sequentially (so they can be selected using a random number generator
    
    # Create two dataframes, one for companies that get successfully evaluated and one for companies which failed during valuation for whatever reason
    winners_df = pd.DataFrame(columns=['Company name', 'Symbol', 'Industry', 'Date', 'Share price', 'Analyst target', 'PE', 'Debt rating', 'Min VPS', 'Avg VPS', 'Max VPS'])
    losers_df = pd.DataFrame(columns=['Company name', 'Symbol', 'Industry', 'Date', 'Share price', 'Analyst target', 'PE', 'Debt rating', 'reason for failure'])
    winners_filename = '/home/dinesh/Documents/Valuations/usa/winners.xlsx'
    losers_filename = '/home/dinesh/Documents/Valuations/usa/losers.xlsx'
    
    win_idx = 0
    lose_idx = 0
    num_stocks = 10
    for i in range(0,num_stocks): # Process 100 random stocks per run of this code
        random_index = randint(0, sel_stocks.shape[0]-1)
        
        symbol = sel_stocks.loc[random_index,'Exchange:Ticker'].split(':')[1]
        industry = sel_stocks.loc[random_index,'Industry Group']
    
        if industry in ['Investments & Asset Management', 'Brokerage & Investment Banking', 'R.E.I.T.', 'Bank (Money Center)']:
            continue # We don't want to value these companies. They don't fit into traditional valuation methods
    
        print('Processing ', symbol)    
        
        sresponse = get_statements(symbol)
        fresponse = get_fundamentals(symbol)
        if fresponse['result'] == 'failure':
            losers_df.loc[lose_idx, 'Company name'] = symbol # Workaround to the problem that we don't have the company name at this stage 
            losers_df.loc[lose_idx, 'Symbol'] = symbol 
            losers_df.loc[lose_idx, 'Industry'] = industry
            losers_df.loc[lose_idx, 'Date'] = date.today()
            losers_df.loc[lose_idx, 'Share price'] = np.nan
            losers_df.loc[lose_idx, 'Analyst target'] = np.nan
            losers_df.loc[lose_idx, 'PE'] = np.nan
            losers_df.loc[lose_idx, 'Debt rating'] = ''
            losers_df.loc[lose_idx, 'reason for failure'] = fresponse['reason for failure']
            lose_idx = lose_idx + 1
            print('Failed to process ', symbol, ' because;')        
            print(fresponse['reason for failure']) 
    
            if lose_idx > 3 or i == (num_stocks-1): # Once we have assembled 10 losers, commit the df to the excel file to prevent unexpected data loss
                print('Committing losers to file')
                if os.path.exists(losers_filename):
                    writer = pd.ExcelWriter(losers_filename, mode='a', if_sheet_exists='overlay', engine = 'openpyxl')
                    losers_df.to_excel(writer, 'Summary', index=False, header=False, startrow = writer.book['Summary'].max_row)
                else:
                    writer = pd.ExcelWriter(losers_filename, engine = 'openpyxl')
                    losers_df.to_excel(writer, 'Summary', index=False, header=True)
                writer.close()
                lose_idx = 0
            continue 
        
        response = value_company(symbol, industry, fresponse['fundamentals'])
        if response['result'] == 'failure' and 'Alpha Vantage' in response['reason for failure']:
            time.sleep(60) # Wait 60 seconds and try one more time
            response = value_company(symbol, industry, fresponse['fundamentals'])
             
        if response['result'] == 'failure':
            losers_df.loc[lose_idx, 'Company name'] = response['Company name']
            losers_df.loc[lose_idx, 'Symbol'] = symbol 
            losers_df.loc[lose_idx, 'Industry'] = industry
            losers_df.loc[lose_idx, 'Date'] = date.today()
            losers_df.loc[lose_idx, 'Share price'] = response['Share price']
            losers_df.loc[lose_idx, 'Analyst target'] = response['Analyst target']
            losers_df.loc[lose_idx, 'PE'] = response['PE']
            losers_df.loc[lose_idx, 'Debt rating'] = response['Debt rating']
            losers_df.loc[lose_idx, 'reason for failure'] = response['reason for failure']
            lose_idx = lose_idx + 1
            print('Failed to process ', symbol, ' because;')        
            print(response['reason for failure']) 
            if lose_idx > 3 or i == (num_stocks-1): # Once we have assembled 10 losers, commit the df to the excel file to prevent unexpected data loss
                print('Committing losers to file')
                if os.path.exists(losers_filename):
                    writer = pd.ExcelWriter(losers_filename, mode='a', if_sheet_exists='overlay', engine = 'openpyxl')
                    losers_df.to_excel(writer, 'Summary', index=False, header=False, startrow = writer.book['Summary'].max_row)
                else:
                    writer = pd.ExcelWriter(losers_filename, engine = 'openpyxl')
                    losers_df.to_excel(writer, 'Summary', index=False, header=True)
                writer.close()
                lose_idx = 0
        
        elif response['result'] == 'success':
            winners_df.loc[win_idx, 'Company name'] = response['Company name']
            winners_df.loc[win_idx, 'Symbol'] = symbol 
            winners_df.loc[win_idx, 'Industry'] = industry
            winners_df.loc[win_idx, 'Date'] = date.today()
            winners_df.loc[win_idx, 'Share price'] = response['Share price']
            winners_df.loc[win_idx, 'Analyst target'] = response['Analyst target']
            winners_df.loc[win_idx, 'PE'] = response['PE']
            winners_df.loc[win_idx, 'Debt rating'] = response['Debt rating']
            winners_df.loc[win_idx, 'Min VPS'] = response['value_df'].min().min()
            winners_df.loc[win_idx, 'Avg VPS'] = response['value_df'].min().min()
            winners_df.loc[win_idx, 'Max VPS'] = response['value_df'].max().max()
            win_idx = win_idx + 1
            print('Successfully processed ', symbol)        
            
            
            if win_idx > 0 or i == (num_stocks-1): # Once we have assembled 10 winners, commit the df to the excel file to prevent unexpected data loss
                print('Committing winners to file')
                if os.path.exists(winners_filename):
                    writer = pd.ExcelWriter(winners_filename, mode='a', if_sheet_exists='overlay', engine = 'openpyxl')
                    winners_df.to_excel(writer, 'Summary', index=False, header=False, startrow = writer.book['Summary'].max_row)
                else:
                    writer = pd.ExcelWriter(winners_filename, engine = 'openpyxl')
                    winners_df.to_excel(writer, 'Summary', index=False, header=True)
                writer.close()
                win_idx = 0
        
        else:
            print('WTF! We should not be here')
    
